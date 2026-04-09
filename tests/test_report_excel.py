import os
from datetime import date
import openpyxl
import pytest
from reports.excel import generate_store_report, generate_raw_backup, REPORT_COLUMNS


def _make_single_month_rows():
    return [
        {"CreatedOn": date(2026, 1, 5), "StoreId": 100, "RegisterName": "Belchicken Aalst", "0%": 0, "6%": 450.20, "12%": 0, "21%": 0, "Bancontact": 150.00, "Cash": 100.20, "Betalen met kaart": 200.00, "UberEats": 0, "TakeAway": 0, "Deliveroo": 0},
        {"CreatedOn": date(2026, 1, 2), "StoreId": 100, "RegisterName": "Belchicken Aalst", "0%": 0, "6%": 500.50, "12%": 0, "21%": 0, "Bancontact": 200.00, "Cash": 100.50, "Betalen met kaart": 200.00, "UberEats": 0, "TakeAway": 0, "Deliveroo": 0},
        {"CreatedOn": date(2026, 1, 4), "StoreId": 100, "RegisterName": "Belchicken Aalst", "0%": 0, "6%": 810.00, "12%": 0, "21%": 0, "Bancontact": 400.00, "Cash": 210.00, "Betalen met kaart": 200.00, "UberEats": 0, "TakeAway": 0, "Deliveroo": 0},
        {"CreatedOn": date(2026, 1, 3), "StoreId": 100, "RegisterName": "Belchicken Aalst", "0%": 0, "6%": 620.30, "12%": 0, "21%": 0, "Bancontact": 300.00, "Cash": 120.30, "Betalen met kaart": 200.00, "UberEats": 0, "TakeAway": 0, "Deliveroo": 0},
    ]


def _make_quarterly_rows():
    return [
        {"CreatedOn": date(2026, 1, 2), "StoreId": 100, "RegisterName": "Belchicken Aalst", "0%": 0, "6%": 500.50, "12%": 0, "21%": 0, "Bancontact": 200.00, "Cash": 100.50, "Betalen met kaart": 200.00, "UberEats": 0, "TakeAway": 0, "Deliveroo": 0},
        {"CreatedOn": date(2026, 1, 3), "StoreId": 100, "RegisterName": "Belchicken Aalst", "0%": 0, "6%": 620.30, "12%": 0, "21%": 0, "Bancontact": 300.00, "Cash": 120.30, "Betalen met kaart": 200.00, "UberEats": 0, "TakeAway": 0, "Deliveroo": 0},
        {"CreatedOn": date(2026, 2, 1), "StoreId": 100, "RegisterName": "Belchicken Aalst", "0%": 0, "6%": 700.00, "12%": 0, "21%": 0, "Bancontact": 350.00, "Cash": 150.00, "Betalen met kaart": 200.00, "UberEats": 0, "TakeAway": 0, "Deliveroo": 0},
        {"CreatedOn": date(2026, 2, 2), "StoreId": 100, "RegisterName": "Belchicken Aalst", "0%": 0, "6%": 550.10, "12%": 0, "21%": 0, "Bancontact": 250.00, "Cash": 100.10, "Betalen met kaart": 200.00, "UberEats": 0, "TakeAway": 0, "Deliveroo": 0},
        {"CreatedOn": date(2026, 3, 1), "StoreId": 100, "RegisterName": "Belchicken Aalst", "0%": 0, "6%": 900.00, "12%": 0, "21%": 0, "Bancontact": 400.00, "Cash": 200.00, "Betalen met kaart": 300.00, "UberEats": 0, "TakeAway": 0, "Deliveroo": 0},
    ]


def test_monthly_report_has_one_sheet(tmp_path):
    path = generate_store_report(_make_single_month_rows(), "January 2026", "Belchicken Aalst", str(tmp_path))
    wb = openpyxl.load_workbook(path)
    assert len(wb.sheetnames) == 1


def test_monthly_report_sheet_named_by_month(tmp_path):
    path = generate_store_report(_make_single_month_rows(), "January 2026", "Belchicken Aalst", str(tmp_path))
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames[0] == "January"


def test_monthly_report_has_correct_columns(tmp_path):
    path = generate_store_report(_make_single_month_rows(), "January 2026", "Belchicken Aalst", str(tmp_path))
    wb = openpyxl.load_workbook(path)
    headers = [cell.value for cell in wb.active[1]]
    assert headers == REPORT_COLUMNS


def test_monthly_report_rows_sorted_by_created_on(tmp_path):
    path = generate_store_report(_make_single_month_rows(), "January 2026", "Belchicken Aalst", str(tmp_path))
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    dates = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert dates == sorted(dates)


def test_quarterly_report_has_three_sheets(tmp_path):
    path = generate_store_report(_make_quarterly_rows(), "Q1 - March 2026", "Belchicken Aalst", str(tmp_path))
    wb = openpyxl.load_workbook(path)
    assert len(wb.sheetnames) == 3


def test_quarterly_report_sheets_named_by_month(tmp_path):
    path = generate_store_report(_make_quarterly_rows(), "Q1 - March 2026", "Belchicken Aalst", str(tmp_path))
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["January", "February", "March"]


def test_quarterly_report_each_sheet_has_correct_month_data(tmp_path):
    path = generate_store_report(_make_quarterly_rows(), "Q1 - March 2026", "Belchicken Aalst", str(tmp_path))
    wb = openpyxl.load_workbook(path)
    assert wb["January"].max_row - 1 == 2
    assert wb["February"].max_row - 1 == 2
    assert wb["March"].max_row - 1 == 1


def test_file_naming_convention(tmp_path):
    path = generate_store_report(_make_single_month_rows(), "January 2026", "Belchicken Aalst", str(tmp_path))
    assert os.path.basename(path) == "January 2026 - VAT Accounting Report - Belchicken Aalst.xlsx"


def test_excel_file_is_valid(tmp_path):
    path = generate_store_report(_make_single_month_rows(), "January 2026", "Belchicken Aalst", str(tmp_path))
    wb = openpyxl.load_workbook(path)
    assert wb is not None


def test_raw_backup_has_all_rows(tmp_path):
    rows = _make_quarterly_rows()
    path = generate_raw_backup(rows, "Q1 - March 2026", str(tmp_path))
    wb = openpyxl.load_workbook(path)
    assert wb.active.max_row - 1 == len(rows)


def test_raw_backup_naming_convention(tmp_path):
    path = generate_raw_backup(_make_quarterly_rows(), "Q1 - March 2026", str(tmp_path))
    assert os.path.basename(path) == "Q1 - March 2026 - VAT Raw Report.xlsx"


def test_raw_backup_has_correct_columns(tmp_path):
    path = generate_raw_backup(_make_quarterly_rows(), "Q1 - March 2026", str(tmp_path))
    wb = openpyxl.load_workbook(path)
    headers = [cell.value for cell in wb.active[1]]
    assert headers == REPORT_COLUMNS
