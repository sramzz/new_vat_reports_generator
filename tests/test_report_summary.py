import os
import openpyxl
from reports.summary import generate_summary

def _make_store_results():
    return [
        {"store_id": 100, "store_name": "Belchicken Aalst", "report_url": "https://drive.google.com/file/d/abc/view"},
        {"store_id": 200, "store_name": "Belchicken Brugge", "report_url": "https://drive.google.com/file/d/def/view"},
        {"store_id": 300, "store_name": "Belchicken Leuven", "report_url": "https://drive.google.com/file/d/ghi/view"},
    ]

def test_summary_has_correct_columns(tmp_path):
    path = generate_summary(_make_store_results(), "Q1 - March 2026", str(tmp_path))
    wb = openpyxl.load_workbook(path)
    headers = [cell.value for cell in wb.active[1]]
    assert headers == ["Store ID", "Store Name", "Report URL"]

def test_summary_has_one_row_per_store(tmp_path):
    path = generate_summary(_make_store_results(), "Q1 - March 2026", str(tmp_path))
    wb = openpyxl.load_workbook(path)
    assert wb.active.max_row - 1 == 3

def test_summary_urls_are_populated(tmp_path):
    path = generate_summary(_make_store_results(), "Q1 - March 2026", str(tmp_path))
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for row_num in range(2, ws.max_row + 1):
        url = ws.cell(row=row_num, column=3).value
        assert url is not None and url.startswith("https://")

def test_summary_file_is_valid_xlsx(tmp_path):
    path = generate_summary(_make_store_results(), "Q1 - March 2026", str(tmp_path))
    wb = openpyxl.load_workbook(path)
    assert wb is not None

def test_summary_naming_convention(tmp_path):
    path = generate_summary(_make_store_results(), "Q1 - March 2026", str(tmp_path))
    assert os.path.basename(path) == "Q1 - March 2026 - VAT Summary Report.xlsx"
