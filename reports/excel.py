import calendar
import logging
import os
from collections import defaultdict
from openpyxl import Workbook

logger = logging.getLogger("vat_reports")

REPORT_COLUMNS = [
    "CreatedOn", "RegisterName", "0%", "6%", "12%", "21%",
    "Bancontact", "Cash", "Betalen met kaart", "UberEats", "TakeAway", "Deliveroo",
]

def _write_sheet(ws, rows: list[dict]) -> None:
    ws.append(REPORT_COLUMNS)
    sorted_rows = sorted(rows, key=lambda r: r["CreatedOn"])
    for row in sorted_rows:
        ws.append([row[col] for col in REPORT_COLUMNS])

def _group_by_month(rows: list[dict]) -> dict[str, list[dict]]:
    groups: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        month_name = calendar.month_name[row["CreatedOn"].month]
        groups[month_name].append(row)
    return dict(groups)

def generate_store_report(rows: list[dict], report_name: str, store_name: str, output_dir: str) -> str:
    filename = f"{report_name} - VAT Accounting Report - {store_name}.xlsx"
    path = os.path.join(output_dir, filename)
    logger.info(f"Generating store report: {filename} ({len(rows)} rows)")
    wb = Workbook()
    months = _group_by_month(rows)
    month_order = list(calendar.month_name)[1:]
    sorted_months = sorted(months.keys(), key=lambda m: month_order.index(m))
    for i, month_name in enumerate(sorted_months):
        if i == 0:
            ws = wb.active
            ws.title = month_name
        else:
            ws = wb.create_sheet(title=month_name)
        _write_sheet(ws, months[month_name])
    wb.save(path)
    logger.info(f"Store report saved: {path} ({len(sorted_months)} sheets)")
    return path

def generate_raw_backup(rows: list[dict], report_name: str, output_dir: str) -> str:
    filename = f"{report_name} - VAT Raw Report.xlsx"
    path = os.path.join(output_dir, filename)
    logger.info(f"Generating raw backup: {filename} ({len(rows)} rows)")
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Data"
    _write_sheet(ws, rows)
    wb.save(path)
    logger.info(f"Raw backup saved: {path}")
    return path
